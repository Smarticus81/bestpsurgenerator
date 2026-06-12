"""Auto-discover and classify input files in the data/input directory.

Phase 1: keyword match on filename.
Phase 2: for unmatched files, peek at content and use AI to classify.
Phase 3: if AI can't classify, ask the user via Prompt.ask().
"""
import logging
from pathlib import Path
from typing import Dict, List, Optional

from rich.console import Console
from rich.table import Table
from rich.prompt import Prompt

from llm_client import create_message
from config import MODEL

logger = logging.getLogger(__name__)
console = Console()

# ── File type classification keywords ──────────────────────────────
# Maps category -> (list of filename substrings, description)
FILE_TYPE_KEYWORDS = {
    "sales":       (["sales", "distribution", "units_sold", "shipment"],
                    "Sales / distribution data"),
    "complaints":  (["complaint", "adverse", "vigilance", "mdr_report"],
                    "Complaint records"),
    "capa":        (["capa", "corrective", "preventive"],
                    "CAPA records"),
    "cer":         (["cer", "clinical_evaluation", "clinical evaluation"],
                    "Clinical Evaluation Report"),
    "ifu":         (["ifu", "instructions_for_use", "instructions for use"],
                    "Instructions for Use"),
    "rmf":         (["rmf", "risk_management", "risk management"],
                    "Risk Management File"),
    "ract":        (["ract", "risk_assessment", "risk assessment", "risk_control"],
                    "Risk Assessment and Control Table"),
    "pms_plan":    (["pms", "post_market_surveillance", "post-market surveillance", "pms_plan", "plan"],
                    "PMS Plan"),
    "pmcf":        (["pmcf", "post_market_clinical", "post-market clinical",
                     "clinical_safety", "clinical safety",
                     "clinical_performance", "clinical performance"],
                    "PMCF Report/Plan"),
    "fsca":        (["fsca", "field_safety", "field safety"],
                    "FSCA data"),
    "external_db": (["maude", "external_db", "external_database", "registry", "eudamed",
                     "external_event", "external event"],
                    "External database search results"),
    "literature":  (["literature", "lit_search", "lit search"],
                    "Scientific literature search results"),
    "previous_psur": (["previous_psur", "prior_psur", "previous psur"],
                      "Previous PSUR"),
    "device_context": (["device_context", "device context"],
                       "Device Context (JSON metadata) — ONLY source of truth for device identity; "
                       "matched by exact filename 'device_context' only"),
    "coding_dictionary": (["coding_dictionary", "imdrf_codes", "imdrf_dictionary",
                           "annex_a", "annex_f", "harm_mdp", "code_dictionary"],
                          "Coding dictionary / taxonomy (e.g. IMDRF Annex A/F)"),
    "chart_sales":    (["sales_chart", "chart_sales", "sales_trend", "volume_chart",
                        "distribution_chart"],
                       "Sales / distribution trend chart (image)"),
    "chart_trend":    (["trend_chart", "chart_trend", "trend_ucl", "ucl_chart",
                        "complaint_chart", "complaint_rate", "rate_chart", "rate_trend"],
                       "Complaint rate / UCL trend chart (image)"),
    "analysis_workbook": (["tables_and_charts", "tables_charts", "analysis_workbook",
                           "analysis", "workbook", "psur_workbook", "psur_data",
                           "consolidated", "unified"],
                          "Pre-computed analysis workbook (Excel with multiple sheets: sales tables, complaint trending, harms, incidents)"),
}


def _peek_file_content(filepath: Path, max_chars: int = 2000) -> str:
    """Peek at the first bit of a file's content for classification."""
    ext = filepath.suffix.lower()
    try:
        if ext in (".csv", ".txt", ".md", ".markdown", ".json"):
            with open(filepath, "r", errors="replace") as fh:
                return fh.read(max_chars)
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = [f"Sheet names: {wb.sheetnames}"]
            ws = wb.active
            if ws:
                for row in ws.iter_rows(max_row=5, values_only=True):
                    parts.append(" | ".join(str(c) for c in row if c is not None))
            wb.close()
            return "\n".join(parts)[:max_chars]
        elif ext == ".docx":
            from docx import Document as DocxDoc
            doc = DocxDoc(filepath)
            text = "\n".join(p.text for p in doc.paragraphs[:20] if p.text.strip())
            return text[:max_chars]
        elif ext == ".pdf":
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:2]:
                    text += (page.extract_text() or "") + "\n"
                    if len(text) >= max_chars:
                        break
                return text[:max_chars]
    except Exception:
        pass
    return ""


def _ai_classify_file(filename: str, snippet: str, categories_desc: str) -> Optional[str]:
    """Use Claude/OpenAI to classify a file based on its name and content snippet."""
    try:
        response = create_message(
            model=MODEL,
            max_tokens=100,
            temperature=0.0,
            messages=[{
                "role": "user",
                "content": f"""Classify this medical device file into exactly ONE category.

Filename: {filename}

Content preview:
{snippet[:1500]}

Categories:
{categories_desc}

For image files (PNG, JPG, etc.), classify based on the filename.
Images whose names suggest a chart or graph should be classified as chart_sales or chart_trend.

Reply with ONLY the category key (e.g., "sales", "complaints", "cer", "chart_sales"). If unsure, reply "unknown"."""
            }]
        )
        result = response.content[0].text.strip().lower().replace('"', '').replace("'", "")
        if result == "unknown":
            return None
        return result
    except Exception:
        return None


def auto_discover_inputs(input_dir: Path) -> Dict[str, List[Path]]:
    """
    Scan input_dir for all files and classify each.

    Phase 1: keyword match on filename.
    Phase 2: for unmatched files, peek at content and use AI to classify.
    Phase 3: if AI can't classify, ask the user via Prompt.ask().

    Every file should be a first-class source — nothing is relegated to 'extra'.
    """
    classified: Dict[str, List[Path]] = {cat: [] for cat in FILE_TYPE_KEYWORDS}
    classified["extra"] = []

    if not input_dir.exists():
        return classified

    supported_exts = {
        ".csv", ".xlsx", ".xls", ".docx", ".doc", ".pdf",
        ".json", ".txt", ".md", ".markdown",
        ".png", ".jpg", ".jpeg", ".tiff", ".bmp",
    }

    unmatched: List[Path] = []

    # Phase 1: keyword match on filename
    for f in sorted(input_dir.iterdir()):
        if f.is_dir() or f.name.startswith(".") or f.name.startswith("~"):
            continue
        if f.suffix.lower() not in supported_exts:
            continue

        stem_lower = f.stem.lower().replace("-", "_")
        matched = False
        for category, (keywords, _desc) in FILE_TYPE_KEYWORDS.items():
            # device_context is strict: exact stem match only ("device_context.json").
            # No fuzzy/substring matches — it is the single source of truth for
            # device identity and must never be inferred from another file.
            if category == "device_context":
                if stem_lower == "device_context":
                    classified[category].append(f)
                    matched = True
                    break
                continue
            if any(kw in stem_lower for kw in keywords):
                classified[category].append(f)
                matched = True
                break
        if not matched:
            unmatched.append(f)

    # Phase 2: content-aware AI classification for unmatched files
    # Images (PNG/JPG/etc.) have no peekable content but the AI can still
    # classify them by filename alone, so we never skip them.
    image_exts = {".png", ".jpg", ".jpeg", ".tiff", ".bmp"}
    if unmatched:
        categories_desc = "\n".join(
            f"  {cat}: {desc}" for cat, (_, desc) in FILE_TYPE_KEYWORDS.items()
        )
        for f in list(unmatched):
            snippet = _peek_file_content(f)
            is_image = f.suffix.lower() in image_exts

            if not snippet and not is_image:
                # Non-image file with no readable content — skip to user prompt
                classified["extra"].append(f)
                unmatched.remove(f)
                continue

            # For images use filename-only hint; for others use content snippet
            if is_image and not snippet:
                snippet = f"(Image file — classify by filename only: {f.name})"

            ai_category = _ai_classify_file(f.name, snippet, categories_desc)
            # device_context is reserved for exact-filename matches in Phase 1
            # only. Block AI from ever assigning it to prevent silent identity
            # corruption (e.g. coding_dictionary.json → device_context).
            if ai_category == "device_context":
                console.print(
                    f"  [yellow]AI tried to classify {f.name} as device_context — "
                    f"blocked. device_context.json must be exact filename match.[/yellow]"
                )
                ai_category = None
            if ai_category and ai_category in FILE_TYPE_KEYWORDS:
                classified[ai_category].append(f)
                unmatched.remove(f)
                console.print(f"  [dim]AI classified {f.name} -> {ai_category}[/dim]")

    # Phase 3: ask user for remaining unmatched files
    if unmatched:
        category_choices = list(FILE_TYPE_KEYWORDS.keys()) + ["extra"]
        for f in unmatched:
            console.print(f"\n  [yellow]Cannot auto-classify: {f.name}[/yellow]")
            choice = Prompt.ask(
                f"  Assign category for [cyan]{f.name}[/cyan]",
                choices=category_choices,
                default="extra",
            )
            classified[choice].append(f)

    return classified


def print_discovered_files(discovered: Dict[str, List[Path]]):
    """Print a summary table of auto-discovered input files."""
    table = Table(title="Auto-Discovered Input Files")
    table.add_column("Category", style="cyan", min_width=16)
    table.add_column("File(s)", style="green")

    total = 0
    for category, files in discovered.items():
        if not files:
            continue
        desc = FILE_TYPE_KEYWORDS.get(category, ([], category))
        label = desc[1] if isinstance(desc, tuple) else category
        names = ", ".join(f.name for f in files)
        table.add_row(label, names)
        total += len(files)

    if total == 0:
        console.print("  [yellow]No input files found in data/input/[/yellow]")
        console.print("  Place your files there and re-run.\n")
    else:
        console.print(table)
        console.print(f"  [dim]{total} file(s) discovered[/dim]\n")
