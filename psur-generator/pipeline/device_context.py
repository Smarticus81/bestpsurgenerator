"""Device metadata extraction from source documents or a pre-authored context file.

Primary path (fast):
  - load_device_context_file() — read a user-supplied device_context.json

Fallback path (LLM):
  - extract_device_context_llm() — auto-detect from file snippets via LLM

Also provides:
- gather_file_snippets() — collect text snippets from discovered input files
- resolve_device_metadata() — validate/default LLM-detected metadata
- build_device_context() — assemble the full device_context dict for orchestrator
"""
import json
import re
import logging
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from llm_client import create_message
from config import MODEL

logger = logging.getLogger(__name__)

# ── Device Context File loader ──────────────────────────────────────


def load_device_context_file(filepath: Path) -> Dict[str, Any]:
    """Load a user-authored device_context.json and map it to internal format.

    Returns a dict with two top-level keys:
      "meta"  — fields consumed by resolve_device_metadata / main.py
      "rich"  — fields merged into build_device_context (device_context dict)

    The caller should use "meta" to skip the LLM auto-detection call
    and "rich" to enrich the device_context before the orchestrator runs.
    """
    with open(filepath, "r", encoding="utf-8") as fh:
        raw: Dict[str, Any] = json.load(fh)

    # --- Parse classification + rule ---
    class_and_rule = raw.get("eu_mdr_classification_and_rule", "")
    device_class = ""
    classification_rule = ""
    if class_and_rule:
        class_and_rule_upper = class_and_rule.strip().upper()
        for candidate in ("CLASS_III", "CLASS_IIB", "CLASS_IIA", "CLASS_I"):
            if candidate.replace("_", " ") in class_and_rule_upper or candidate in class_and_rule_upper:
                device_class = candidate
                break
        # Extract rule portion (e.g. "Rule 6")
        rule_match = re.search(r"(?:rule|annex\s*viii)\s*\d+", class_and_rule, re.IGNORECASE)
        if rule_match:
            classification_rule = rule_match.group(0)

    # --- Single-use / reusable ---
    single_use_raw = raw.get("single_use_or_reusable", "").strip().lower()
    is_reusable = single_use_raw in ("reusable", "multi-use", "multi use", "true")

    # --- Trade names ---
    trade_names = raw.get("device_trade_names", [])
    device_name = trade_names[0] if trade_names else ""

    # --- Notified Body ---
    nb_raw = raw.get("notified_body_name_and_id", "")
    nb_name, nb_number = "", ""
    if nb_raw:
        nb_match = re.search(r"\(?\s*(\d{4})\s*\)?", nb_raw)
        if nb_match:
            nb_number = nb_match.group(1)
            nb_name = nb_raw[:nb_match.start()].strip().rstrip(",- ")
        else:
            nb_name = nb_raw

    # --- Device lifetime ---
    lifetime = raw.get("device_lifetime", {})

    # --- Document references ---
    pms_doc = raw.get("pms_plan_document", {})
    pmcf_doc = raw.get("pmcf_plan_document", {})
    ifu_doc = raw.get("ifu_document", {})
    cer_doc_number = raw.get("cer_document_number_and_version", "")
    cer_date = raw.get("cer_date_or_last_update", "")

    # Build the "meta" dict (same keys as resolve_device_metadata output)
    meta = {
        "device_name": device_name,
        "auto_detected_name": device_name,
        "device_class": device_class or "CLASS_IIB",
        "is_reusable": is_reusable,
        "certificate_number": raw.get("certificate_number", ""),
        "certificate_date": raw.get("certificate_date", ""),
        "psur_cadence": (
            "ANNUALLY" if (device_class or "CLASS_IIB") in ("CLASS_IIB", "CLASS_III")
            else "EVERY_TWO_YEARS"
        ),
        "infocard_number": "",
    }

    # Build the "rich" dict — merged into device_context in build_device_context()
    rich: Dict[str, Any] = {
        # Core descriptive fields
        "device_trade_names": trade_names,
        "device_description": raw.get("device_description", ""),
        "intended_use": raw.get("intended_purpose", ""),
        "indications": raw.get("indications_for_use", []),
        "contraindications": raw.get("contraindications", []),
        "target_patient_population": raw.get("target_patient_population", ""),
        "intended_user_profile": raw.get("intended_user_profile", ""),
        "sterility_status": raw.get("sterility_status", ""),
        "single_use_or_reusable": raw.get("single_use_or_reusable", ""),
        "market_history": raw.get("market_history", ""),
        "device_lifetime": lifetime,
        "manufacturer_info": {
            "company_name": raw.get("manufacturer_name", ""),
            "address_lines": raw.get("manufacturer_address_lines", []),
            "manufacturer_srn": raw.get("manufacturer_srn", ""),
        },
        "authorized_representative_info": {
            "name": (raw.get("authorized_representative") or {}).get("name", ""),
            "address_lines": (raw.get("authorized_representative") or {}).get("address_lines", []),
            "srn": (
                (raw.get("authorized_representative") or {}).get("authorized_representative_srn", "")
                or (raw.get("authorized_representative") or {}).get("srn", "")
            ),
        },
        "uk_mdr_classification_and_rule": raw.get("uk_mdr_classification_and_rule", ""),
        "uk_responsible_person": raw.get("uk_responsible_person", ""),
        "ukca_marking_status": raw.get("ukca_marking_status", ""),
        # Identifiers
        "known_identifiers": {
            "basic_udi_di": raw.get("basic_udi_di_or_device_family_name", ""),
            "model_numbers": raw.get("model_or_catalog_numbers", []),
            "catalog_numbers": raw.get("model_or_catalog_numbers", []),
            "emdn_code": raw.get("emdn_code", ""),
            "gmdn_codes": [raw.get("gmdn_code", "")] if raw.get("gmdn_code") else [],
            "classification_rule_mdr_annex_viii": classification_rule,
            "first_ce_marking_date": raw.get("date_of_first_ce_marking_or_doc", ""),
            "first_declaration_of_conformity_date": raw.get("date_of_first_ce_marking_or_doc", ""),
            "risk_management_file_number": raw.get("risk_management_file_document_number", ""),
            "eu_technical_documentation_number": raw.get("eu_technical_documentation_number", ""),
            "certificate_number": raw.get("certificate_number", ""),
            "certificate_date": raw.get("certificate_date", ""),
            "us_fda_classification": raw.get("us_fda_classification", ""),
            "us_pre_market_submission_number": raw.get("us_pre_market_submission_number", ""),
            "fda_clearance": raw.get("fda_clearance", ""),
            "first_ec_eu_certificate_date": raw.get("certificate_date", ""),
        },
        # Notified body
        "notified_body": {"name": nb_name, "number": nb_number},
        # Document references
        "cer_document": {"number": cer_doc_number, "date": cer_date},
        "pms_plan_document": pms_doc,
        "pmcf_plan_document": pmcf_doc,
        "ifu_document": ifu_doc,
        "other_associated_documents": raw.get("other_associated_documents", []),
    }

    return {"meta": meta, "rich": rich, "raw": raw}


# Snippet size limits per input type (label → max_chars)
_SNIPPET_SOURCES: List[Tuple[str, int]] = [
    ("cer", 4000), ("previous_psur", 3000), ("pms_plan", 3000),
    ("ract", 2000), ("rmf", 2000), ("complaints", 2000),
    ("sales", 2000), ("capa", 2000), ("ifu", 2000),
    ("pmcf", 2000), ("literature", 2000), ("fsca", 2000), ("external_db", 2000),
]


def peek_document_snippet(filepath: Path, max_chars: int = 4000) -> str:
    """Extract a text snippet from a document for auto-detection."""
    ext = filepath.suffix.lower()
    try:
        if ext in (".csv", ".txt", ".md", ".markdown", ".json"):
            with open(filepath, "r", errors="replace") as fh:
                return fh.read(max_chars)
        elif ext in (".xlsx", ".xls"):
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            parts = [f"Sheet names: {wb.sheetnames}"]
            ws = wb.active
            if ws:
                for row in ws.iter_rows(max_row=20, values_only=True):
                    parts.append(" | ".join(str(c) for c in row if c is not None))
            wb.close()
            return "\n".join(parts)[:max_chars]
        elif ext == ".docx":
            from docx import Document as DocxDoc
            doc = DocxDoc(filepath)
            text = "\n".join(p.text for p in doc.paragraphs[:60] if p.text.strip())
            return text[:max_chars]
        elif ext == ".pdf":
            import pdfplumber
            with pdfplumber.open(filepath) as pdf:
                text = ""
                for page in pdf.pages[:4]:
                    text += (page.extract_text() or "") + "\n"
                    if len(text) >= max_chars:
                        break
                return text[:max_chars]
    except Exception:
        pass
    return ""


def extract_device_context_llm(
    file_snippets: Dict[str, str],
) -> Dict[str, Any]:
    """Use a single LLM call to auto-detect device metadata from ALL source documents.

    Accepts a dict of {label: snippet_text} from every discovered input file.
    Extracts: device_name, eu_mdr_classification, is_reusable, denominator_type,
    certificate_number, certificate_date, psur_cadence, infocard_number.
    """

    snippets = []
    # Prioritise high-value documents first, but include everything
    priority_order = ["cer", "previous_psur", "pms_plan", "ract", "rmf",
                      "complaints", "sales", "capa", "ifu", "pmcf", "fsca", "external_db"]
    seen = set()
    for label in priority_order:
        if label in file_snippets and file_snippets[label]:
            snippets.append(f"=== {label.upper()} ===\n{file_snippets[label]}")
            seen.add(label)
    for label, text in file_snippets.items():
        if label not in seen and text:
            snippets.append(f"=== {label.upper()} ===\n{text}")

    if not snippets:
        return {}

    # Truncate combined text to ~15k chars to stay within token budget
    combined = "\n\n".join(snippets)
    if len(combined) > 15000:
        combined = combined[:15000] + "\n... [truncated]"

    prompt = f"""You are a medical-device regulatory analyst. From the source-document excerpts below, extract the following device metadata. Return ONLY a JSON object with these keys:

{{
  "device_name": "<the EXACT commercial/branded product name>",
  "eu_mdr_classification": "<one of: CLASS_I, CLASS_IIA, CLASS_IIB, CLASS_III>",
  "is_reusable": <true if the device is reusable / multi-use, false if single-use / disposable>,
  "denominator_type": "<'units distributed' for single-use, 'procedures' for reusable>",
  "certificate_number": "<EU certificate number if found, else empty string>",
  "certificate_date": "<ISO 8601 date of certificate if found, else empty string>",
  "psur_cadence": "<ANNUALLY or EVERY_TWO_YEARS — based on MDR class>",
  "infocard_number": "<if found, else empty string>"
}}

RULES:
- device_name: extract the REAL branded/trade product name from the documents.
  Look for it in: CER title, sales product columns, complaint device name fields,
  PSUR cover page, PMS plan title, RACT device name, file names, page headers.
  Examples of good names: "Endosee® System", "ALLY Uterine Positioning System".
  Do NOT invent a name — it must appear verbatim in the source text.
  If multiple product names appear, choose the primary subject of the documents.
- is_reusable: look for "single-use", "disposable", "reusable", "multi-use".
- For Class IIb and III devices, psur_cadence should be "ANNUALLY".
- If a field cannot be determined from the text, use empty string for strings or false for booleans.

SOURCE DOCUMENTS:
{combined}

Respond with ONLY the JSON object, no markdown fences or explanation."""

    try:
        response = create_message(
            model=MODEL,
            max_tokens=500,
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.content[0].text.strip()
        # Strip markdown fences if present
        text = re.sub(r'^```(?:json)?\s*', '', text)
        text = re.sub(r'\s*```$', '', text)
        result = json.loads(text)
        return result
    except Exception as e:
        logger.warning(f"LLM device auto-detection failed: {e}")
        return {}


# ── Snippet gathering ───────────────────────────────────────────────


def gather_file_snippets(
    input_paths: Dict[str, Optional[Path]],
    extra_paths: Optional[List[Path]] = None,
) -> Dict[str, str]:
    """Gather text snippets from all discovered input files for LLM device detection.

    Args:
        input_paths: {label: Path_or_None} for each known input category.
        extra_paths: Additional files (e.g. user-supplied extras).

    Returns:
        {label: snippet_text} for every file that produced a readable snippet.
    """
    snippets: Dict[str, str] = {}
    for label, max_chars in _SNIPPET_SOURCES:
        fpath = input_paths.get(label)
        if fpath and fpath.exists():
            snip = peek_document_snippet(fpath, max_chars=max_chars)
            if snip:
                snippets[label] = snip
    for ef in (extra_paths or []):
        p = Path(ef)
        if p.exists():
            snip = peek_document_snippet(p, max_chars=1500)
            if snip:
                snippets[f"extra_{p.stem}"] = snip
    return snippets


# ── Metadata resolution ─────────────────────────────────────────────


def resolve_device_metadata(
    llm_detected: Dict[str, Any],
    cli_device_name: str,
    start_date: str,
    end_date: str,
) -> Dict[str, Any]:
    """Validate / default LLM-detected device metadata and merge with CLI overrides.

    Returns dict with keys:
        device_name, auto_detected_name, device_class, is_reusable,
        certificate_number, certificate_date, psur_cadence, infocard_number.
    """
    detected_name = llm_detected.get("device_name", "")
    device_name = cli_device_name or detected_name

    device_class = llm_detected.get("eu_mdr_classification", "CLASS_IIB")
    if isinstance(device_class, str):
        device_class = device_class.strip().upper()
    if device_class not in {"CLASS_IIA", "CLASS_IIB", "CLASS_III"}:
        device_class = "CLASS_IIB"

    is_reusable = llm_detected.get("is_reusable", False)
    certificate_number = llm_detected.get("certificate_number", "")
    certificate_date = llm_detected.get("certificate_date", "")

    psur_cadence = llm_detected.get("psur_cadence", "")
    if isinstance(psur_cadence, str):
        psur_cadence = psur_cadence.strip().upper()
    if psur_cadence not in {"ANNUALLY", "EVERY_TWO_YEARS"}:
        psur_cadence = (
            "ANNUALLY" if device_class in {"CLASS_IIB", "CLASS_III"} else "EVERY_TWO_YEARS"
        )

    # Period-aware cadence override: if period > 14 months it's biennial
    try:
        _s = datetime.strptime(start_date, "%Y-%m-%d")
        _e = datetime.strptime(end_date, "%Y-%m-%d")
        _period_months = (_e.year - _s.year) * 12 + (_e.month - _s.month) + 1
        if _period_months > 14:
            psur_cadence = "EVERY_TWO_YEARS"
    except (ValueError, TypeError):
        pass

    infocard_number = llm_detected.get("infocard_number", "")

    return {
        "device_name": device_name,
        "auto_detected_name": detected_name,
        "device_class": device_class,
        "is_reusable": is_reusable,
        "certificate_number": certificate_number,
        "certificate_date": certificate_date,
        "psur_cadence": psur_cadence,
        "infocard_number": infocard_number,
    }


# ── Full device-context assembly ────────────────────────────────────


def build_device_context(
    *,
    device_name: str,
    device_class: str,
    is_reusable: bool,
    certificate_number: str,
    certificate_date: str,
    psur_cadence: str,
    infocard_number: str,
    denominator_type: str,
    denominator_description: str,
    parsed_data: Dict[str, Any],
    expanded_context: Dict[str, str],
    input_paths: Dict[str, Optional[Path]],
    context_file_rich: Optional[Dict[str, Any]] = None,
) -> Tuple[Dict[str, Any], str]:
    """Assemble the complete device_context dict consumed by the orchestrator.

    Enriches base metadata with CER extraction data, manufacturer info,
    known identifiers, and previous-PSUR fallbacks.

    Returns:
        (device_context, possibly_updated_device_name)
    """
    device_context: Dict[str, Any] = {
        "device_name": device_name,
        "eu_mdr_classification": device_class,
        "certificate_number": certificate_number,
        "certificate_date": certificate_date,
        "psur_cadence": psur_cadence,
        "infocard_number": infocard_number,
        "is_reusable": is_reusable,
        "denominator_type": denominator_type,
        "denominator_description": denominator_description,
    }

    manufacturer_info: Dict[str, Any] = {
        "company_name": "",
        "address_lines": [],
        "manufacturer_srn": "",
    }
    ar_info: Dict[str, Any] = {}
    notified_body_info: Dict[str, str] = {"name": "", "number": ""}

    known_identifiers: Dict[str, Any] = {
        "basic_udi_di": "",
        "eu_technical_documentation_number": "",
        "us_pre_market_submission_number": "",
        "fda_clearance": "",
        "emdn_code": "",
        "risk_management_file_number": "",
        "classification_rule_mdr_annex_viii": "",
        "model_numbers": [],
        "catalog_numbers": [],
        "gmdn_codes": [],
        "certificates": [],
        "first_ce_marking_date": "",
        "first_declaration_of_conformity_date": "",
        "first_ec_eu_certificate_date": "",
    }

    # ── Device context file enrichment (highest priority) ─────

    context_file_has_mfr = False
    if context_file_rich:
        # Merge descriptive fields
        for key in (
            "device_description", "intended_use", "indications",
            "intended_purpose",
            "contraindications", "target_patient_population",
            "intended_user_profile", "sterility_status",
            "single_use_or_reusable", "market_history", "device_lifetime",
            "device_trade_names", "cer_document", "pms_plan_document",
            "pmcf_plan_document", "ifu_document", "other_associated_documents",
            "eu_mdr_classification_and_rule", "uk_mdr_classification_and_rule",
        ):
            val = context_file_rich.get(key)
            if val:  # non-empty string, non-empty list, non-empty dict
                device_context[key] = val

        cf_mfr = context_file_rich.get("manufacturer_info") or {}
        if context_file_rich.get("manufacturer_name") or cf_mfr.get("company_name"):
            manufacturer_info["company_name"] = context_file_rich.get("manufacturer_name") or cf_mfr.get("company_name", "")
        if context_file_rich.get("manufacturer_address_lines") or cf_mfr.get("address_lines"):
            manufacturer_info["address_lines"] = context_file_rich.get("manufacturer_address_lines") or cf_mfr.get("address_lines", [])
        if context_file_rich.get("manufacturer_srn") or cf_mfr.get("manufacturer_srn"):
            manufacturer_info["manufacturer_srn"] = context_file_rich.get("manufacturer_srn") or cf_mfr.get("manufacturer_srn", "")
        context_file_has_mfr = bool(manufacturer_info.get("company_name"))

        if context_file_rich.get("authorized_representative") or context_file_rich.get("authorized_representative_info"):
            ar_src = context_file_rich.get("authorized_representative") or context_file_rich.get("authorized_representative_info")
            if isinstance(ar_src, dict):
                ar_info = {
                    "name": ar_src.get("name", ""),
                    "address_lines": ar_src.get("address_lines", []),
                    "srn": ar_src.get("authorized_representative_srn") or ar_src.get("srn", ""),
                }

        # Merge identifiers (fill gaps, don't overwrite non-empty)
        cf_ids = context_file_rich.get("known_identifiers", {})
        for k, v in cf_ids.items():
            if v and not known_identifiers.get(k):
                known_identifiers[k] = v

        # Merge notified body
        cf_nb = context_file_rich.get("notified_body", {})
        if cf_nb.get("name") and not notified_body_info.get("name"):
            notified_body_info["name"] = cf_nb["name"]
        if cf_nb.get("number") and not notified_body_info.get("number"):
            notified_body_info["number"] = cf_nb["number"]
        if context_file_rich.get("notified_body_name_and_id") and not notified_body_info.get("name"):
            nb_text = str(context_file_rich["notified_body_name_and_id"])
            notified_body_info["name"] = nb_text
            import re
            m = re.search(r"\b(\d{4})\b", nb_text)
            if m:
                notified_body_info["number"] = m.group(1)

        top_level_id_map = {
            "basic_udi_di_or_device_family_name": "basic_udi_di",
            "emdn_code": "emdn_code",
            "risk_management_file_document_number": "risk_management_file_number",
            "eu_technical_documentation_number": "eu_technical_documentation_number",
            "classification_rule_mdr_annex_viii": "classification_rule_mdr_annex_viii",
            "certificate_number": "certificate_number",
            "certificate_date": "certificate_date",
        }
        for src, dest in top_level_id_map.items():
            val = context_file_rich.get(src)
            if val and dest in known_identifiers and not known_identifiers.get(dest):
                known_identifiers[dest] = val
            elif val and dest in ("certificate_number", "certificate_date") and not device_context.get(dest):
                device_context[dest] = val
        if context_file_rich.get("model_or_catalog_numbers") and not known_identifiers.get("model_numbers"):
            known_identifiers["model_numbers"] = context_file_rich["model_or_catalog_numbers"]
            known_identifiers["catalog_numbers"] = context_file_rich["model_or_catalog_numbers"]

        logger.info("Merged device_context.json rich fields into device_context")

    # ── CER enrichment ──────────────────────────────────────────

    if "cer" in parsed_data and isinstance(parsed_data["cer"], dict):
        cer_data = parsed_data["cer"]
        device_context.update({
            "device_description": cer_data.get("device_description", ""),
            "intended_use": cer_data.get("intended_use", ""),
            "indications": cer_data.get("indications", ""),
            "contraindications": cer_data.get("contraindications", ""),
        })

        # Auto-deduce device name from CER
        cer_device_name = ""
        if cer_data.get("device_detail"):
            cer_device_name = cer_data["device_detail"].get("device_name", "")
        if cer_device_name and (
            not device_name or device_name.startswith("TD") or device_name == "Unknown"
        ):
            device_name = cer_device_name
            device_context["device_name"] = device_name
            logger.info("Auto-deduced device name from CER: %s", device_name)

        # Auto-deduce manufacturer from CER
        mfr_info = cer_data.get("manufacturer_info", {})
        if mfr_info and mfr_info.get("company_name") and not context_file_has_mfr:
            manufacturer_info["company_name"] = mfr_info["company_name"]
            manufacturer_info["address_lines"] = mfr_info.get("address_lines", [])
            manufacturer_info["manufacturer_srn"] = mfr_info.get("manufacturer_srn", "")
            logger.info("Auto-deduced manufacturer from CER: %s", mfr_info["company_name"])

            if mfr_info.get("authorized_representative_name"):
                ar_info = {
                    "name": mfr_info["authorized_representative_name"],
                    "address_lines": mfr_info.get("authorized_representative_address", []),
                    "srn": mfr_info.get("authorized_representative_srn", ""),
                }

        # Regulatory info
        if cer_data.get("regulatory_info"):
            device_context["regulatory_info"] = cer_data["regulatory_info"]
            reg = cer_data["regulatory_info"]

            if reg.get("fda_clearance"):
                known_identifiers["fda_clearance"] = reg["fda_clearance"]
                known_identifiers["us_pre_market_submission_number"] = reg["fda_clearance"]
            if reg.get("certificates"):
                known_identifiers["certificates"] = reg["certificates"]
                if not device_context.get("certificate_number"):
                    certs = reg["certificates"]
                    if isinstance(certs, list) and certs:
                        first_cert = (
                            certs[0] if isinstance(certs[0], dict)
                            else {"number": str(certs[0])}
                        )
                        device_context["certificate_number"] = first_cert.get(
                            "number", str(certs[0])
                        )
                        if first_cert.get("date") and not device_context.get("certificate_date"):
                            device_context["certificate_date"] = first_cert["date"]
                    elif isinstance(certs, str):
                        device_context["certificate_number"] = certs
            if reg.get("classification_eu"):
                known_identifiers["classification_rule_mdr_annex_viii"] = ""
            if reg.get("classification_rule"):
                known_identifiers["classification_rule_mdr_annex_viii"] = reg["classification_rule"]
            if reg.get("eu_technical_documentation_number"):
                known_identifiers["eu_technical_documentation_number"] = reg[
                    "eu_technical_documentation_number"
                ]
            if reg.get("risk_management_file_number"):
                known_identifiers["risk_management_file_number"] = reg[
                    "risk_management_file_number"
                ]
            if reg.get("first_ce_marking_date"):
                known_identifiers["first_ce_marking_date"] = reg["first_ce_marking_date"]
            if reg.get("first_declaration_of_conformity_date"):
                known_identifiers["first_declaration_of_conformity_date"] = reg[
                    "first_declaration_of_conformity_date"
                ]
            # First EC/EU certificate date
            if (
                reg.get("certificates")
                and isinstance(reg["certificates"], list)
                and reg["certificates"]
            ):
                first_cert = reg["certificates"][0]
                if isinstance(first_cert, dict) and first_cert.get("date"):
                    known_identifiers["first_ec_eu_certificate_date"] = first_cert["date"]
            if (
                not known_identifiers.get("first_ec_eu_certificate_date")
                and reg.get("first_ce_marking_date")
            ):
                known_identifiers["first_ec_eu_certificate_date"] = reg["first_ce_marking_date"]
            # Notified body
            if reg.get("notified_body_name"):
                notified_body_info["name"] = reg["notified_body_name"]
            if reg.get("notified_body_number"):
                notified_body_info["number"] = reg["notified_body_number"]

        # Device identifiers
        if cer_data.get("device_identifiers"):
            device_context["device_identifiers"] = cer_data["device_identifiers"]
            ids = cer_data["device_identifiers"]
            if ids.get("udi_di"):
                known_identifiers["basic_udi_di"] = ids["udi_di"]
            if ids.get("emdn_codes"):
                known_identifiers["emdn_code"] = (
                    ids["emdn_codes"][0].get("code", "") if ids["emdn_codes"] else ""
                )
            if ids.get("model_numbers"):
                known_identifiers["model_numbers"] = ids["model_numbers"]
            if ids.get("catalog_numbers"):
                known_identifiers["catalog_numbers"] = ids["catalog_numbers"]
            if ids.get("gmdn_codes"):
                known_identifiers["gmdn_codes"] = ids["gmdn_codes"]

        # Extra CER detail blocks
        for key in (
            "device_detail", "indications_detail", "population_info",
            "ifu_info", "safety_efficacy_detail",
        ):
            if cer_data.get(key):
                device_context[key] = cer_data[key]

    # ── Finalize metadata ───────────────────────────────────────

    device_context["known_identifiers"] = known_identifiers
    device_context["manufacturer_info"] = manufacturer_info
    device_context["authorized_representative_info"] = ar_info
    device_context["notified_body"] = notified_body_info

    # Previous-PSUR fallback for manufacturer / device name
    if not manufacturer_info.get("company_name") and isinstance(
        parsed_data.get("previous_psur"), dict
    ):
        prev = parsed_data["previous_psur"]
        if prev.get("manufacturer"):
            manufacturer_info["company_name"] = prev["manufacturer"]
            device_context["manufacturer_info"] = manufacturer_info
            logger.info(
                "Auto-deduced manufacturer from previous PSUR: %s", prev["manufacturer"]
            )
        if prev.get("device_name") and (not device_name or device_name.startswith("TD")):
            device_name = prev["device_name"]
            device_context["device_name"] = device_name
            logger.info(
                "Auto-deduced device name from previous PSUR: %s", device_name
            )

    if expanded_context:
        device_context["expanded_context"] = expanded_context

    # Available inputs list so agents know what data was provided
    available_inputs = []
    for label in (
        "sales", "complaints", "capa", "cer", "ifu", "rmf", "ract",
        "pms_plan", "pmcf", "literature", "fsca", "external_db", "previous_psur",
    ):
        path = input_paths.get(label)
        if path and path.exists():
            available_inputs.append(label)
    device_context["available_inputs"] = available_inputs

    return device_context, device_name
