"""Parser for pre-computed analysis workbooks (Excel).

Extracts structured data from workbooks containing pre-built PSUR tables
and chart data.  Recognised sheet names (case-insensitive):

- **sales_tables**       → Annual sales by region (Table 1) + monthly units
- **complaint_trending**  → Monthly complaint rates, mean, UCL, breach flag
- **harms_table**         → Table 7 (harm × MDP cross-tab with rates/occurrence)
- **section_d**           → Tables 2/3/4 (serious incidents by IMDRF code)

The parser returns a dict with keys that can be merged directly into the
pipeline's ``parsed_data``, ``statistics``, and ``chart_data`` structures.
"""
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

logger = logging.getLogger(__name__)


# ───────────────────────────────────────────────────────────────────
# Public entry point
# ───────────────────────────────────────────────────────────────────

def parse_analysis_workbook(path: Path) -> Dict[str, Any]:
    """Parse a pre-computed analysis workbook.

    Returns a dict with optional keys:
        sales_by_region    – list of {region, …year columns…, pct_of_total}
        monthly_sales      – dict YYYY-MM → units (float)
        total_units        – int/float worldwide total
        complaint_trending – list of {month, complaints, cumulative_sales,
                             complaint_rate, mean_rate, ucl, breach}
        trend_summary      – {mean_rate, ucl, data_points, monthly_rates,
                             monthly_labels, monthly_complaints, monthly_sales}
        harms_table        – list of {harm, mdp, current_rate, current_count,
                             cumulative_rate, cumulative_count,
                             max_expected_code, max_expected_rate_numeric}
        section_d_table2   – list of {region, imdrf_code, imdrf_term,
                             count, rate_pct, complaint_numbers}
        section_d_table3   – list of dicts (same shape)
        section_d_table4   – list of dicts (health impact)
    """
    path = Path(path)
    if not path.exists():
        logger.warning("Analysis workbook not found: %s", path)
        return {}

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    sheet_map = {s.lower().replace(" ", "_"): s for s in wb.sheetnames}

    result: Dict[str, Any] = {}

    # ── Sales tables ───────────────────────────────────────────────
    # NOTE: "sales" alone is NOT a candidate here — it matches the raw
    # sales-data sheet, not the pre-computed sales analysis tables.
    sales_key = _find_sheet(sheet_map, ["sales_tables", "sales_table",
                                         "sales_analysis", "table_1"])
    if sales_key:
        ws = wb[sheet_map[sales_key]]
        rows = _read_rows(ws)
        region, monthly = _parse_sales_tables(rows)
        if region:
            result["sales_by_region"] = region
        if monthly:
            result["monthly_sales"] = monthly
            result["total_units"] = int(round(sum(monthly.values())))

    # ── Complaint trending ─────────────────────────────────────────
    trend_key = _find_sheet(sheet_map, ["complaint_trending", "complaint_rate",
                                        "trending", "trend", "complaint_trend",
                                        "complaint_rates"])
    if trend_key:
        ws = wb[sheet_map[trend_key]]
        rows = _read_rows(ws)
        trending, summary = _parse_complaint_trending(rows)
        if trending:
            result["complaint_trending"] = trending
        if summary:
            result["trend_summary"] = summary

    # ── Harms table ────────────────────────────────────────────────
    harms_key = _find_sheet(sheet_map, ["harms_table", "harms", "table_7",
                                        "table7", "harm_table", "harms_tables"])
    if harms_key:
        ws = wb[sheet_map[harms_key]]
        rows = _read_rows(ws)
        harms = _parse_harms_table(rows)
        if harms:
            # Auto-compute Grand Total if not present
            if not any(h.get("is_grand_total") for h in harms):
                harm_level = [h for h in harms if not h.get("mdp")]
                gt_cur_count = sum(h["current_count"] for h in harm_level)
                gt_cum_count = sum(h["cumulative_count"] for h in harm_level)
                gt_cur_rate = sum(h["current_rate"] for h in harm_level)
                gt_cum_rate = sum(h["cumulative_rate"] for h in harm_level)
                harms.append({
                    "harm": "Grand Total",
                    "mdp": "",
                    "current_rate": round(gt_cur_rate, 6),
                    "current_count": gt_cur_count,
                    "cumulative_rate": round(gt_cum_rate, 6),
                    "cumulative_count": gt_cum_count,
                    "max_expected_code": "",
                    "max_expected_rate_numeric": 0.0,
                    "is_grand_total": True,
                })
                logger.info("Harms table: auto-computed Grand Total (%d current, %d cumulative)",
                            gt_cur_count, gt_cum_count)
            result["harms_table"] = harms

    # ── Section D (serious incidents) ──────────────────────────────
    sec_d_key = _find_sheet(sheet_map, ["section_d", "section_d_-_incidents",
                                        "serious_incidents", "incidents",
                                        "section_d_incidents"])
    if sec_d_key:
        ws = wb[sheet_map[sec_d_key]]
        rows = _read_rows(ws)
        t2, t3, t4 = _parse_section_d(rows)
        if t2:
            result["section_d_table2"] = t2
        if t3:
            result["section_d_table3"] = t3
        if t4:
            result["section_d_table4"] = t4

    # ── External DB Search ─────────────────────────────────────────
    ext_db_key = _find_sheet(sheet_map, ["external_db_search", "external_db",
                                         "external_database", "registry_search",
                                         "maude", "table_10"])
    if ext_db_key:
        ws = wb[sheet_map[ext_db_key]]
        rows = _read_rows(ws)
        ext_db = _parse_external_db_search(rows)
        if ext_db:
            result["external_db"] = ext_db

    # ── Detect raw data sheets (for unified workbook mode) ─────────
    raw_sheets: List[str] = []
    raw_sheet_names: Dict[str, str] = {}
    for category, candidates in [
        ("sales", ["sales", "sales_data", "sales_raw"]),
        ("complaints", ["complaints", "complaints_data", "complaint_data"]),
        ("capa", ["capa", "capa_data", "capa_raw"]),
    ]:
        found = _find_sheet(sheet_map, candidates)
        if found:
            raw_sheets.append(category)
            raw_sheet_names[category] = sheet_map[found]

    if raw_sheets:
        result["raw_sheets_available"] = raw_sheets
        result["raw_sheet_names"] = raw_sheet_names

    # ── Config sheet (device metadata & cover page) ────────────────
    config_key = _find_sheet(sheet_map, ["config", "configuration",
                                          "settings", "setup"])
    if config_key:
        ws = wb[sheet_map[config_key]]
        rows = _read_rows(ws)
        config = _parse_config_sheet(rows)
        if config:
            result["config"] = config

    # ── Section H – FSCA ───────────────────────────────────────────
    fsca_key = _find_sheet(sheet_map, ["section_h_-_fsca", "section_h",
                                        "fsca", "field_safety"])
    if fsca_key:
        ws = wb[sheet_map[fsca_key]]
        rows = _read_rows(ws)
        fsca_rows, fsca_summary = _parse_section_h_fsca(rows)
        if fsca_rows is not None:
            result["fsca_table"] = fsca_rows
        if fsca_summary:
            result["fsca_summary"] = fsca_summary

    # ── Section I – CAPA ───────────────────────────────────────────
    capa_key = _find_sheet(sheet_map, ["section_i_-_capa", "section_i",
                                        "capa_table"])
    if capa_key:
        ws = wb[sheet_map[capa_key]]
        rows = _read_rows(ws)
        capa_rows, capa_summary = _parse_section_i_capa(rows)
        if capa_rows is not None:
            result["capa_section_i"] = capa_rows
        if capa_summary:
            result["capa_summary"] = capa_summary

    wb.close()

    logger.info("Analysis workbook parsed: %s — keys=%s",
                path.name, list(result.keys()))
    return result


# ───────────────────────────────────────────────────────────────────
# Helpers
# ───────────────────────────────────────────────────────────────────

def _find_sheet(sheet_map: Dict[str, str], candidates: List[str]) -> Optional[str]:
    """Return the first matching normalised sheet name, or None."""
    for c in candidates:
        if c in sheet_map:
            return c
    return None


def _read_rows(ws) -> List[List[Any]]:
    """Read all rows from a worksheet as lists of raw values."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _str(val) -> str:
    """Stringify a cell value, collapsing None to ''."""
    if val is None:
        return ""
    return str(val).strip()


def _num(val, default=0) -> float:
    """Coerce a cell value to float, stripping text like parenthesised counts."""
    if val is None:
        return default
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s:
        return default
    # Strip percentage signs
    s = s.replace("%", "").strip()
    try:
        return float(s)
    except ValueError:
        return default


def _parse_rate_and_count(cell_text: str) -> Tuple[float, int]:
    """Parse a cell like '0.0008 (582)' into (rate, count)."""
    if not cell_text or not str(cell_text).strip():
        return (0.0, 0)
    s = str(cell_text).strip()
    # Pattern: rate (count)
    m = re.match(r'([\d.]+)\s*\((\d+)\)', s)
    if m:
        return (float(m.group(1)), int(m.group(2)))
    # Just a number
    try:
        return (float(s), 0)
    except ValueError:
        return (0.0, 0)


def _parse_month(val) -> Optional[str]:
    """Convert a cell value to YYYY-MM string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m")
    s = str(val).strip()
    # Already YYYY-MM
    if re.match(r'^\d{4}-\d{2}$', s):
        return s
    # YYYY-MM-DD or datetime string
    m = re.match(r'^(\d{4}-\d{2})', s)
    if m:
        return m.group(1)
    # "Apr 2022" etc.
    for fmt in ("%b %Y", "%B %Y", "%b-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m")
        except ValueError:
            continue
    return None


# ───────────────────────────────────────────────────────────────────
# Sheet parsers
# ───────────────────────────────────────────────────────────────────

def _parse_sales_tables(rows: List[List[Any]]) -> Tuple[List[Dict], Dict[str, float]]:
    """Parse the sales_tables sheet.

    Supports both standard layout (data at col A) and offset layout
    (data at col B, as in the user's surveillance workbook where col A
    is empty/reserved).

    Returns (region_rows, monthly_sales).
    """
    region_rows: List[Dict] = []
    monthly_sales: Dict[str, float] = {}

    if not rows:
        return region_rows, monthly_sales

    # --- Annual by-region section ---
    # Find the header row containing "Region" (scan all columns)
    region_header_idx = None
    region_col_offset = 0
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if _str(cell).lower() == "region":
                region_header_idx = i
                region_col_offset = j
                break
        if region_header_idx is not None:
            break

    if region_header_idx is not None:
        hdr_row = rows[region_header_idx]
        headers = [_str(hdr_row[c]) if c < len(hdr_row) else ""
                   for c in range(region_col_offset, len(hdr_row))]
        # Read data rows until footnote marker or blank stretch
        for r in rows[region_header_idx + 1:]:
            region_name = _str(r[region_col_offset]) if region_col_offset < len(r) else ""
            if not region_name:
                continue
            # Stop at footnote markers
            if region_name.startswith("¹") or region_name.startswith("²") or region_name.startswith("["):
                break
            entry: Dict[str, Any] = {"region": region_name}
            for col_idx in range(1, len(headers)):
                abs_col = region_col_offset + col_idx
                if abs_col >= len(r):
                    continue
                col_name = headers[col_idx]
                if not col_name:
                    continue
                val = r[abs_col]
                if "%" in col_name.lower() or "total" in col_name.lower():
                    entry["pct_of_total"] = _num(val)
                else:
                    entry[col_name] = _num(val)
            region_rows.append(entry)

    # --- Monthly section ---
    # Find "Month" + "Units Sold" header (scan all columns)
    monthly_header_idx = None
    monthly_col_offset = 0
    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            if _str(cell).lower() == "month" and i > 0:
                # Check context: next col says "Units Sold" or prev row says "Monthly"
                next_col = _str(row[j + 1]).lower() if j + 1 < len(row) else ""
                prev_row_cell = _str(rows[i - 1][j]).lower() if j < len(rows[i - 1]) else ""
                if "unit" in next_col or "monthly" in prev_row_cell:
                    monthly_header_idx = i
                    monthly_col_offset = j
                    break
        if monthly_header_idx is not None:
            break

    if monthly_header_idx is not None:
        for r in rows[monthly_header_idx + 1:]:
            month_val = r[monthly_col_offset] if monthly_col_offset < len(r) else None
            units_val = _num(r[monthly_col_offset + 1]) if monthly_col_offset + 1 < len(r) else 0
            month_str = _parse_month(month_val)
            if month_str and units_val > 0:
                monthly_sales[month_str] = units_val
            elif not _str(month_val):
                if monthly_sales:
                    break  # blank row after data = end of section

    return region_rows, monthly_sales


def _parse_complaint_trending(rows: List[List[Any]]) -> Tuple[List[Dict], Dict[str, Any]]:
    """Parse the complaint_trending sheet.

    Returns (trending_rows, summary_dict).
    """
    trending: List[Dict] = []
    summary: Dict[str, Any] = {}

    if not rows or len(rows) < 2:
        return trending, summary

    # First row is headers; data starts at row 1
    monthly_rates: List[float] = []
    monthly_labels: List[str] = []
    monthly_complaints: List[int] = []
    monthly_sales_list: List[float] = []
    mean_rate = 0.0
    ucl = 0.0

    for r in rows[1:]:
        month_str = _parse_month(r[0])
        if month_str is None:
            continue

        complaints = int(_num(r[1])) if len(r) > 1 else 0
        cum_sales = _num(r[2]) if len(r) > 2 else 0
        rate = _num(r[3]) if len(r) > 3 else 0.0
        m_rate = _num(r[4]) if len(r) > 4 else 0.0
        u = _num(r[5]) if len(r) > 5 else 0.0
        breach_val = _str(r[6]) if len(r) > 6 else ""
        breach = breach_val.upper() in ("YES", "TRUE", "1", "BREACH")

        trending.append({
            "month": month_str,
            "complaints": complaints,
            "cumulative_sales": cum_sales,
            "complaint_rate": rate,
            "mean_rate": m_rate,
            "ucl": u,
            "breach": breach,
        })

        monthly_rates.append(rate)
        monthly_labels.append(month_str)
        monthly_complaints.append(complaints)
        monthly_sales_list.append(cum_sales)
        mean_rate = m_rate
        ucl = u

    if monthly_rates:
        summary = {
            "mean_rate": mean_rate,
            "ucl": ucl,
            "data_points": len(monthly_rates),
            "monthly_rates": monthly_rates,
            "monthly_labels": monthly_labels,
            "monthly_complaints": monthly_complaints,
            "monthly_sales": monthly_sales_list,
        }

    return trending, summary


def _parse_harms_table(rows: List[List[Any]]) -> List[Dict]:
    """Parse the harms_table sheet (Table 7 cross-tab).

    Supports two layouts:
    - **Standard**: data starts at column A (col index 0).
    - **Offset** (user workbook): data starts at column F (col index 5)
      because columns A-E hold date-range metadata.

    The parser auto-detects the offset by finding which column contains
    the "Harm" header.  It then reads the **first** table it encounters
    (12-month or annual) and skips any duplicated biennial table below.

    Returns list of dicts with harm, mdp, rates, counts, max expected.
    """
    result: List[Dict] = []
    if not rows:
        return result

    # ── Detect data-start column and header row ────────────────────
    header_idx = None
    col_offset = 0          # 0-based column where data starts

    for i, row in enumerate(rows):
        for j, cell in enumerate(row):
            cell_text = _str(cell).lower()
            # Must be a SHORT cell that says "harm" — not a long title
            # string that happens to contain "harm" somewhere inside.
            if cell_text == "harm" or (cell_text.startswith("harm") and len(cell_text) < 25):
                # Verify the same row has neighbouring context columns
                rest = [_str(c).lower() for c in row[j:]]
                if any("current" in c or "medical device problem" in c or "mdp" in c for c in rest):
                    header_idx = i
                    col_offset = j
                    break
        if header_idx is not None:
            break

    if header_idx is None:
        return result

    # ── Find the sub-header row (contains "Medical Device Problem") ─
    # and determine where actual data rows begin.
    data_start = header_idx + 1
    for peek in range(header_idx + 1, min(header_idx + 4, len(rows))):
        peek_text = _str(rows[peek][col_offset]) if peek < len(rows) and col_offset < len(rows[peek]) else ""
        lower = peek_text.lower()
        # skip sub-header rows (e.g. "   Medical Device Problem", date ranges
        # like "Jul-2024 to Jun-2025").  Use word-boundary " to " to avoid
        # false positives on words like "symptoms".
        is_date_range = " to " in lower or re.search(r'\d{4}\s+to\s+\d{4}', lower) is not None
        if "medical device problem" in lower or is_date_range or not peek_text:
            data_start = peek + 1
        else:
            break

    # ── Parse data rows ────────────────────────────────────────────
    current_harm = ""
    for r in rows[data_start:]:
        # Stop if we hit a second "Table 7" header block (biennial
        # duplicate), a blank separator, or another "Harm" header.
        raw_col0 = str(r[col_offset]) if col_offset < len(r) and r[col_offset] is not None else ""
        col0 = raw_col0.strip()

        if not col0:
            # A blank row between the two tables → stop
            if result:
                break
            continue

        lower_col0 = col0.lower()

        # Stop markers: second header block, footnotes, or "Table 7" title
        if "table 7" in lower_col0 or (lower_col0.startswith("[") and "]" in lower_col0):
            if result:
                break
            continue

        # Data columns relative to col_offset
        val1 = r[col_offset + 1] if col_offset + 1 < len(r) else ""
        val2 = r[col_offset + 2] if col_offset + 2 < len(r) else ""
        val3 = r[col_offset + 3] if col_offset + 3 < len(r) else ""
        val4 = r[col_offset + 4] if col_offset + 4 < len(r) else ""

        if "grand total" in lower_col0:
            cur_rate, cur_count = _parse_rate_and_count(val1)
            cum_rate, cum_count = _parse_rate_and_count(val2)
            result.append({
                "harm": "Grand Total",
                "mdp": "",
                "current_rate": cur_rate,
                "current_count": cur_count,
                "cumulative_rate": cum_rate,
                "cumulative_count": cum_count,
                "max_expected_code": "",
                "max_expected_rate_numeric": 0.0,
                "is_grand_total": True,
            })
            break  # Grand total is always last in the first table

        # Indented rows (starting with spaces) are MDP sub-rows
        is_mdp = raw_col0.startswith("   ") or raw_col0.startswith("\t")
        cur_rate, cur_count = _parse_rate_and_count(val1)
        cum_rate, cum_count = _parse_rate_and_count(val2)
        max_code = _str(val3)
        max_rate_num = _num(val4)

        if is_mdp:
            result.append({
                "harm": current_harm,
                "mdp": col0,
                "current_rate": cur_rate,
                "current_count": cur_count,
                "cumulative_rate": cum_rate,
                "cumulative_count": cum_count,
                "max_expected_code": max_code,
                "max_expected_rate_numeric": max_rate_num,
                "is_grand_total": False,
            })
        else:
            current_harm = col0
            result.append({
                "harm": current_harm,
                "mdp": "",  # Harm-level aggregate
                "current_rate": cur_rate,
                "current_count": cur_count,
                "cumulative_rate": cum_rate,
                "cumulative_count": cum_count,
                "max_expected_code": max_code,
                "max_expected_rate_numeric": max_rate_num,
                "is_grand_total": False,
            })

    return result


def _parse_section_d(rows: List[List[Any]]) -> Tuple[List[Dict], List[Dict], List[Dict]]:
    """Parse section_d sheet containing Tables 2, 3, and 4.

    Returns (table2, table3, table4).
    """
    table2: List[Dict] = []
    table3: List[Dict] = []
    table4: List[Dict] = []

    if not rows:
        return table2, table3, table4

    # Split rows into table segments based on "Table X" markers
    segments: List[Tuple[str, int, int]] = []
    marker_indices: List[Tuple[str, int]] = []

    for i, row in enumerate(rows):
        text = _str(row[0]).lower()
        if "table 2" in text:
            marker_indices.append(("table2", i))
        elif "table 3" in text:
            marker_indices.append(("table3", i))
        elif "table 4" in text:
            marker_indices.append(("table4", i))

    # Build segments
    for idx, (name, start) in enumerate(marker_indices):
        end = marker_indices[idx + 1][1] if idx + 1 < len(marker_indices) else len(rows)
        segments.append((name, start, end))

    for name, start, end in segments:
        segment_rows = rows[start:end]
        if name in ("table2", "table3"):
            parsed = _parse_incident_table(segment_rows)
            if name == "table2":
                table2 = parsed
            else:
                table3 = parsed
        elif name == "table4":
            table4 = _parse_health_impact_table(segment_rows)

    return table2, table3, table4


def _parse_incident_table(segment_rows: List[List[Any]]) -> List[Dict]:
    """Parse Tables 2 or 3 (Serious Incidents by IMDRF code)."""
    result: List[Dict] = []

    # Find header row with "Region" and "IMDRF"
    header_idx = None
    for i, row in enumerate(segment_rows):
        cells = [_str(c).lower() for c in row]
        if any("region" in c for c in cells) and any("imdrf" in c for c in cells):
            header_idx = i
            break

    if header_idx is None:
        return result

    current_region = ""
    for r in segment_rows[header_idx + 1:]:
        region_cell = _str(r[0])
        imdrf_cell = _str(r[1]) if len(r) > 1 else ""
        count_val = int(_num(r[2])) if len(r) > 2 else 0
        rate_str = _str(r[3]) if len(r) > 3 else ""
        complaint_nums = _str(r[4]) if len(r) > 4 else ""

        if region_cell:
            current_region = region_cell

        if not imdrf_cell and not count_val:
            continue

        # Parse "A0401 – Break" into code and term
        imdrf_code = ""
        imdrf_term = imdrf_cell
        m = re.match(r'^([A-Z]\d+(?:\.\d+)?)\s*[–—-]\s*(.+)$', imdrf_cell)
        if m:
            imdrf_code = m.group(1)
            imdrf_term = m.group(2).strip()

        # Check if this is a TOTAL row
        if "total" in imdrf_cell.lower():
            result.append({
                "region": current_region,
                "imdrf_code": "TOTAL",
                "imdrf_term": "TOTAL",
                "count": count_val,
                "rate_pct": rate_str,
                "complaint_numbers": "",
                "is_total": True,
            })
            continue

        result.append({
            "region": current_region,
            "imdrf_code": imdrf_code,
            "imdrf_term": imdrf_term,
            "count": count_val,
            "rate_pct": rate_str,
            "complaint_numbers": complaint_nums,
            "is_total": False,
        })

    return result


def _parse_health_impact_table(segment_rows: List[List[Any]]) -> List[Dict]:
    """Parse Table 4 (IMDRF Annex F Health Impact)."""
    result: List[Dict] = []

    # Find the actual column header row (has both "Health Impact" and "Region")
    header_idx = None
    for i, row in enumerate(segment_rows):
        cells = [_str(c).lower() for c in row]
        has_impact = any("health impact" in c for c in cells)
        has_region = any("region" in c for c in cells)
        if has_impact and has_region:
            header_idx = i
            break
    # Fallback: find first row with "Region" after any "annex f" / title row
    if header_idx is None:
        for i, row in enumerate(segment_rows):
            cells = [_str(c).lower() for c in row]
            if any("region" in c for c in cells) and any("incident" in c or "no." in c for c in cells):
                header_idx = i
                break

    if header_idx is None:
        return result

    headers = [_str(c) for c in segment_rows[header_idx]]

    for r in segment_rows[header_idx + 1:]:
        impact = _str(r[0])
        if not impact:
            continue

        entry: Dict[str, Any] = {"health_impact": impact}
        for col_idx in range(1, len(r)):
            if col_idx < len(headers):
                col_name = headers[col_idx]
                if col_name:
                    entry[col_name] = _str(r[col_idx]) if col_idx < len(r) else ""
        result.append(entry)

    return result


# ───────────────────────────────────────────────────────────────────
# External DB Search (Table 10)
# ───────────────────────────────────────────────────────────────────

# Maps common header text to the Table 10 schema field names.
_EXT_DB_HEADER_MAP = {
    "database": "database_registry",
    "database_registry": "database_registry",
    "registry": "database_registry",
    "database/registry": "database_registry",
    "search_date": "search_date",
    "date": "search_date",
    "search_terms": "search_terms",
    "terms": "search_terms",
    "total_matches": "total_matches",
    "matches": "total_matches",
    "relevant_findings": "relevant_findings",
    "findings": "relevant_findings",
    "benchmark_vs_similar_devices": "benchmark_vs_similar_devices",
    "benchmark": "benchmark_vs_similar_devices",
    "regulatory_actions_affecting_similar_devices": "regulatory_actions_affecting_similar_devices",
    "regulatory_actions": "regulatory_actions_affecting_similar_devices",
    "rmf_update_reference": "rmf_update_reference",
    "rmf_update": "rmf_update_reference",
    "rmf_reference": "rmf_update_reference",
}


def _parse_external_db_search(rows: List[List[Any]]) -> List[Dict]:
    """Parse the external_db_search sheet into Table 10 schema rows.

    Each row represents one database/registry that was searched, with
    its results and relevance assessment.
    """
    result: List[Dict] = []
    if not rows or len(rows) < 2:
        return result

    # Normalise headers
    raw_headers = [_str(c).lower().replace(" ", "_") for c in rows[0]]
    headers = [_EXT_DB_HEADER_MAP.get(h, h) for h in raw_headers]

    for row in rows[1:]:
        # Skip blank rows
        if not any(_str(c) for c in row):
            continue

        entry: Dict[str, Any] = {}
        for i, header in enumerate(headers):
            if i < len(row):
                val = row[i]
                if header == "total_matches":
                    entry[header] = int(_num(val))
                else:
                    entry[header] = _str(val)

        # Ensure all Table 10 schema fields are present
        entry.setdefault("database_registry", "N/A")
        entry.setdefault("total_matches", 0)
        entry.setdefault("relevant_findings", "")
        entry.setdefault("benchmark_vs_similar_devices", "N/A")
        entry.setdefault("regulatory_actions_affecting_similar_devices", "N/A")
        entry.setdefault("rmf_update_reference", "N/A")

        result.append(entry)

    return result


# ───────────────────────────────────────────────────────────────────
# Config sheet parser
# ───────────────────────────────────────────────────────────────────

# Maps row labels (lower-cased) → output dict key.  The value column
# is always col C (index 2).
_CONFIG_ROW_MAP = {
    "product / device name":    "device_name",
    "product/device name":      "device_name",
    "device name":              "device_name",
    "surveillance start date":  "surveillance_start",
    "surveillance end date":    "surveillance_end",
    "update frequency":         "psur_cadence",
    "reporting frequency":      "psur_cadence",
    "historical data start date": "historical_start",
    "historical data end date": "historical_end",
    "company name":             "manufacturer_name",
    "manufacturer name":        "manufacturer_name",
    "company address":          "manufacturer_address",
    "manufacturer address":     "manufacturer_address",
    "manufacturer srn":         "manufacturer_srn",
    "authorized representative": "authorized_rep",
    "authorised representative": "authorized_rep",
    "ar srn":                   "ar_srn",
    "notified body name":       "notified_body_name",
    "notified body number":     "notified_body_number",
    "certificate number":       "certificate_number",
    "basic udi-di":             "basic_udi_di",
    "udi-di":                   "basic_udi_di",
    "device trade name":        "device_trade_name",
    "emdn code":                "emdn_code",
    "gmdn code":                "gmdn_code",
    "device risk class":        "device_class",
    "psur cadence":             "psur_cadence",
    "data collection period":   "data_collection_period",
}


def _parse_config_sheet(rows: List[List[Any]]) -> Dict[str, Any]:
    """Parse the Config sheet, extracting device metadata into a flat dict.

    The layout is:
      Col A: section headers / empty
      Col B: Setting label (e.g. "Product / Device Name")
      Col C: Value
      Col D: Notes (ignored)

    We scan every row for a known label in col B and grab the value
    from col C.
    """
    result: Dict[str, Any] = {}
    if not rows:
        return result

    for row in rows:
        if len(row) < 3:
            continue
        label = _str(row[1]).lower().strip()
        key = _CONFIG_ROW_MAP.get(label)
        if key is None:
            continue
        raw = row[2]
        if raw is None:
            continue
        # Convert dates to ISO strings
        if isinstance(raw, datetime):
            result[key] = raw.strftime("%Y-%m-%d")
        elif isinstance(raw, (int, float)) and key in ("notified_body_number", "gmdn_code"):
            result[key] = int(raw)
        else:
            result[key] = str(raw).strip()

    # Normalise psur_cadence to canonical form
    cadence = result.get("psur_cadence", "")
    if isinstance(cadence, str):
        cadence_lower = cadence.lower()
        if "biennial" in cadence_lower or "2" in cadence_lower:
            result["psur_cadence"] = "biennial"
        elif "annual" in cadence_lower or "1" in cadence_lower:
            result["psur_cadence"] = "annual"

    logger.info("Config sheet: %d fields extracted", len(result))
    return result


# ───────────────────────────────────────────────────────────────────
# Section H – FSCA parser
# ───────────────────────────────────────────────────────────────────

def _parse_section_h_fsca(rows: List[List[Any]]) -> Tuple[List[Dict], Dict[str, Any]]:
    """Parse section H (Table 8 – FSCA) from the workbook.

    Returns (fsca_rows, fsca_summary).
    - fsca_rows: list of dicts with action_type, ref_number, issuing_date,
      scope, status, rationale, regions, mhra_date
    - fsca_summary: {total, open, closed}
    """
    fsca_rows: List[Dict] = []
    summary: Dict[str, Any] = {}

    if not rows:
        return fsca_rows, summary

    # Find the header row (contains "Type of Action" or "Manufacturer Ref")
    header_idx = None
    for i, row in enumerate(rows):
        cells_lower = [_str(c).lower() for c in row]
        if any("type of action" in c for c in cells_lower) or (
            any("manufacturer" in c for c in cells_lower) and any("ref" in c for c in cells_lower)
        ):
            header_idx = i
            break

    if header_idx is not None:
        # Read data rows between header and summary
        for r in rows[header_idx + 1:]:
            # Stop at "Summary" marker
            first_cell = _str(r[0]).lower()
            if "summary" in first_cell or "total" in first_cell:
                break

            # Skip blank rows
            if not any(_str(c) for c in r[:8]):
                continue

            issuing_raw = r[2] if len(r) > 2 else None
            mhra_raw = r[7] if len(r) > 7 else None

            fsca_rows.append({
                "action_type": _str(r[0]),
                "ref_number": _str(r[1]) if len(r) > 1 else "",
                "issuing_date": issuing_raw.strftime("%Y-%m-%d") if isinstance(issuing_raw, datetime) else _str(issuing_raw),
                "scope": _str(r[3]) if len(r) > 3 else "",
                "status": _str(r[4]) if len(r) > 4 else "",
                "rationale": _str(r[5]) if len(r) > 5 else "",
                "regions": _str(r[6]) if len(r) > 6 else "",
                "mhra_date": mhra_raw.strftime("%Y-%m-%d") if isinstance(mhra_raw, datetime) else _str(mhra_raw),
            })

    # Parse summary section
    for row in rows:
        label = _str(row[0]).lower()
        if "total fsca" in label:
            summary["total"] = int(_num(row[1]) if len(row) > 1 else 0)
        elif "open fsca" in label:
            summary["open"] = int(_num(row[1]) if len(row) > 1 else 0)
        elif "closed fsca" in label:
            summary["closed"] = int(_num(row[1]) if len(row) > 1 else 0)

    logger.info("Section H FSCA: %d rows, summary=%s", len(fsca_rows), summary)
    return fsca_rows, summary


# ───────────────────────────────────────────────────────────────────
# Section I – CAPA parser
# ───────────────────────────────────────────────────────────────────

def _parse_section_i_capa(rows: List[List[Any]]) -> Tuple[List[Dict], Dict[str, Any]]:
    """Parse section I (Table 9 – CAPA) from the workbook.

    Returns (capa_rows, capa_summary).
    - capa_rows: list of dicts with capa_id, date_initiated, source,
      description, action_taken, status, target_close_date
    - capa_summary: {total, open, closed, overdue}
    """
    capa_rows: List[Dict] = []
    summary: Dict[str, Any] = {}

    if not rows:
        return capa_rows, summary

    # Find the header row (contains "CAPA ID")
    header_idx = None
    for i, row in enumerate(rows):
        cells_lower = [_str(c).lower() for c in row]
        if any("capa id" in c for c in cells_lower) or (
            any("capa" in c for c in cells_lower) and any("date" in c for c in cells_lower)
        ):
            header_idx = i
            break

    if header_idx is not None:
        for r in rows[header_idx + 1:]:
            first_cell = _str(r[0]).lower()
            if "summary" in first_cell or "total" in first_cell:
                break

            if not any(_str(c) for c in r[:7]):
                continue

            date_init = r[1] if len(r) > 1 else None
            target_close = r[6] if len(r) > 6 else None

            capa_rows.append({
                "capa_id": _str(r[0]),
                "date_initiated": date_init.strftime("%Y-%m-%d") if isinstance(date_init, datetime) else _str(date_init),
                "source": _str(r[2]) if len(r) > 2 else "",
                "description": _str(r[3]) if len(r) > 3 else "",
                "action_taken": _str(r[4]) if len(r) > 4 else "",
                "status": _str(r[5]) if len(r) > 5 else "",
                "target_close_date": target_close.strftime("%Y-%m-%d") if isinstance(target_close, datetime) else _str(target_close),
            })

    # Parse summary section
    for row in rows:
        label = _str(row[0]).lower()
        if "total capa" in label:
            summary["total"] = int(_num(row[1]) if len(row) > 1 else 0)
        elif "open capa" in label:
            summary["open"] = int(_num(row[1]) if len(row) > 1 else 0)
        elif "closed capa" in label:
            summary["closed"] = int(_num(row[1]) if len(row) > 1 else 0)
        elif "overdue" in label:
            summary["overdue"] = int(_num(row[1]) if len(row) > 1 else 0)

    logger.info("Section I CAPA: %d rows, summary=%s", len(capa_rows), summary)
    return capa_rows, summary
